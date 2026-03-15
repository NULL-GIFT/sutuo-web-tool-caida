import os
import re
import pandas as pd
import openpyxl
import zipfile
import shutil
import warnings
import math
import tempfile
import streamlit as st

warnings.filterwarnings("ignore")


# ================= 核心处理函数 (保留了你的原始逻辑) =================
def normalize_id(value):
    if value is None: return None
    s = str(value).strip()
    if s.endswith('.0'): s = s[:-2]
    match = re.search(r'(\d{8,})', s)
    if match: return match.group(1)
    return None


def clean_filename(filename):
    try:
        return filename.encode('cp437').decode('gbk')
    except:
        return filename


def recursive_search_files(directory, temp_dir_root):
    excel_files = []
    valid_extensions = ('.xlsx', '.xls', '.xlsm')
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            if file.lower().endswith(valid_extensions) and not file.startswith('~$'):
                excel_files.append(file_path)
            elif file.lower().endswith('.zip'):
                try:
                    with zipfile.ZipFile(file_path, 'r') as zip_ref:
                        extract_subdir = os.path.join(temp_dir_root, f"unzip_{file}_{os.urandom(4).hex()}")
                        os.makedirs(extract_subdir, exist_ok=True)
                        for member in zip_ref.infolist():
                            member.filename = clean_filename(member.filename)
                            zip_ref.extract(member, path=extract_subdir)
                        excel_files.extend(recursive_search_files(extract_subdir, temp_dir_root))
                except:
                    pass
    return excel_files


def extract_info(filepath):
    filename = os.path.basename(filepath)
    student_id = None
    final_score = 0
    try:
        fname_id = normalize_id(filename)
        if fname_id:
            student_id = fname_id
        else:
            df_header = pd.read_excel(filepath, header=None, nrows=6)
            for col in df_header.columns:
                for cell in df_header[col].astype(str):
                    extracted = normalize_id(cell)
                    if extracted and "..." not in cell and len(set(extracted)) > 3:
                        student_id = extracted
                        break
                if student_id: break

        if not student_id: return None, 0

        df_search = pd.read_excel(filepath, header=None, nrows=10)
        header_row = None
        for i, row in df_search.iterrows():
            if "应加分数" in row.astype(str).values:
                header_row = i
                break

        if header_row is None: return student_id, 0

        df = pd.read_excel(filepath, header=header_row)
        target_col = None
        for col in df.columns:
            if "应加分数" in str(col):
                target_col = col
                break

        scores = []
        stop_words = ['加分项目', '扣分项目', '总综评分', '签名', '小计']
        for _, row in df.iterrows():
            if any(w in str(row.iloc[0]) for w in stop_words): break
            try:
                val = float(row[target_col])
                if not math.isnan(val): scores.append(val)
            except:
                continue

        final_score = max(scores) if scores else 0
        return student_id, final_score
    except Exception as e:
        return None, 0


# 把原来的 process_all 改造成接收参数的版本
def process_data(personal_dir, class_dir, output_dir, temp_extract_dir):
    log_messages = []

    def write_log(msg):
        log_messages.append(msg)

    write_log("========== 开始全能兼容模式 ==========")
    write_log("🚀 正在扫描个人表（支持 xls/xlsx/zip）...")
    all_files = recursive_search_files(personal_dir, temp_extract_dir)
    db = {}

    write_log(f"📄 共发现 {len(all_files)} 个文件。")

    for f in all_files:
        sid, score = extract_info(f)
        if sid:
            if sid in db:
                db[sid] = max(db[sid], score)
            else:
                db[sid] = score

    write_log(f"✅ 提取完毕，有效学号数据库: {len(db)} 人。\n")

    class_files = [f for f in os.listdir(class_dir) if f.endswith('.xlsx')]
    total_filled = 0
    output_files = []

    for c_file in class_files:
        write_log(f"📘 正在写入总表: {c_file}")
        c_path = os.path.join(class_dir, c_file)
        save_path = os.path.join(output_dir, f"已处理_{c_file}")

        try:
            wb = openpyxl.load_workbook(c_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header_row, id_col, score_col, name_col = None, None, None, None

                for r in range(1, 10):
                    row_vals = [str(c.value).strip() for c in ws[r] if c.value]
                    if any(x in row_vals for x in ['学号', '学 号']) and any(x in row_vals for x in ['组织任职', '组织认职']):
                        header_row = r
                        for idx, cell in enumerate(ws[r], 1):
                            v = str(cell.value).strip()
                            if v in ['学号', '学 号']: id_col = idx
                            if v in ['组织任职', '组织认职']: score_col = idx
                            if v in ['姓名', '姓 名']: name_col = idx
                        break

                if not header_row: continue

                filled_list = []
                for r in range(header_row + 1, ws.max_row + 1):
                    cell_val = ws.cell(row=r, column=id_col).value
                    sid = normalize_id(cell_val)
                    if sid and sid in db:
                        score = db[sid]
                        ws.cell(row=r, column=score_col).value = score
                        name = "未知姓名"
                        if name_col:
                            cell_name = ws.cell(row=r, column=name_col).value
                            if cell_name: name = str(cell_name).strip()
                        filled_list.append(f"{name}({score})")
                        total_filled += 1

                if filled_list:
                    write_log(f"   └── [{sheet_name}] 录入 {len(filled_list)} 人: " + "，".join(filled_list))

            wb.save(save_path)
            output_files.append(save_path)
            write_log(f"💾 保存成功。\n")
        except Exception as e:
            write_log(f"❌ 出错: {e}")

    write_log("=" * 50)
    write_log(f"🎉 全部搞定！录入 {total_filled} 人。")

    return output_files, "\n".join(log_messages)


# ================= Streamlit 网页前端界面 =================
st.set_page_config(page_title="综测自动核录系统", page_icon="📊")

st.title("📊 社区综测自动核录系统")
st.markdown("上传个人的综测证明材料（支持打包成 ZIP）和班级总表模板，系统将自动提取分数并填入总表。")

# --- 1. 文件上传区 ---
st.subheader("📁 第一步：上传文件")
col1, col2 = st.columns(2)

with col1:
    personal_uploads = st.file_uploader("1. 上传个人证明材料",
                                        type=["xlsx", "xls", "zip"],
                                        accept_multiple_files=True,
                                        help="支持多选 Excel 文件，或直接把整个文件夹打包成 1 个 ZIP 上传。")

with col2:
    class_upload = st.file_uploader("2. 上传总综评分模板",
                                    type=["xlsx"],
                                    accept_multiple_files=False,
                                    help="只需要传一个包含班级名单的 .xlsx 模板")

# --- 初始化系统的“记忆” ---
if 'processed' not in st.session_state:
    st.session_state.processed = False
    st.session_state.excel_data = None
    st.session_state.excel_name = None
    st.session_state.log_text = None

# --- 2. 核心处理区 ---
st.subheader("⚙️ 第二步：执行录入")
if st.button("🚀 开始提取并填表", type="primary"):
    if not personal_uploads:
        st.error("❌ 请先上传个人证明材料！")
    elif not class_upload:
        st.error("❌ 请先上传总表模板！")
    else:
        with st.spinner("后台疯狂处理中... 请稍候..."):
            # 创建临时沙箱目录
            with tempfile.TemporaryDirectory() as base_temp_dir:
                p_dir = os.path.join(base_temp_dir, "personal")
                c_dir = os.path.join(base_temp_dir, "class")
                out_dir = os.path.join(base_temp_dir, "output")
                ex_dir = os.path.join(base_temp_dir, "extract")

                os.makedirs(p_dir); os.makedirs(c_dir)
                os.makedirs(out_dir); os.makedirs(ex_dir)

                # 把用户传进来的内存文件，保存到沙箱里
                for p_file in personal_uploads:
                    with open(os.path.join(p_dir, p_file.name), "wb") as f:
                        f.write(p_file.getbuffer())

                with open(os.path.join(c_dir, class_upload.name), "wb") as f:
                    f.write(class_upload.getbuffer())

                # 调用你的核心程序
                results, log_text = process_data(p_dir, c_dir, out_dir, ex_dir)

                if results:
                    # ✅ 核心改动：把处理结果存进系统的“记忆”里！
                    st.session_state.processed = True
                    st.session_state.log_text = log_text
                    st.session_state.excel_name = os.path.basename(results[0])
                    # 把生成的文件读取成二进制数据存起来
                    with open(results[0], "rb") as f:
                        st.session_state.excel_data = f.read()
                else:
                    st.warning("⚠️ 处理结束，但没有生成新的总表，请检查控制台日志或模板格式。")
                    st.session_state.processed = False
                    st.session_state.log_text = log_text

# --- 3. 结果下载与日志显示 (独立出来，靠记忆判断) ---
# 只要记忆里显示处理过了，就把这块区域展示出来，点一万次下载也不会消失
if st.session_state.processed:
    st.success("✅ 处理完成！请在下方下载结果。")
    st.subheader("⬇️ 第三步：下载结果")

    # 使用两列布局，把两个下载按钮并排放在一起
    col_dl1, col_dl2 = st.columns(2)

    with col_dl1:
        st.download_button(
            label="📥 下载填写完毕的总表",
            data=st.session_state.excel_data,
            file_name=st.session_state.excel_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with col_dl2:
        st.download_button(
            label="📝 下载最终核对清单 (.txt)",
            data=st.session_state.log_text,
            file_name="最终核对清单.txt",
            mime="text/plain"
        )

    st.text_area("📄 运行日志 (详细核对清单)", value=st.session_state.log_text, height=300)
